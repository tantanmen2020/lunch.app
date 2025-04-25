# モジュールのインポート
import random, PIL.Image, PIL.ImageTk, threading, webbrowser, os
import tkinter as tk
from PIL import Image, ImageTk
from playsound import playsound
from openpyxl import Workbook, load_workbook
from datetime import datetime

# 店名とスタンプの数を定義
food_stamps = {
    'カレー': 0,
    'ラーメン': 0,
    'コンビニ': 0,
    'スーパー': 0,
    'うどん平': 0,
    'マカロニキッチン': 0,
    'キッチングローリ': 0,
    '九一郎': 0,
    '食処 やま利': 0,
    '元祖ビックリ亭': 0,
    'かどや食堂': 0,
    'みかちゃん家のいろどりごはん': 0,
    'ロイヤルホスト': 0,
    'キッシュとパンのジョー': 0,
    'Home Kitchen COTALO': 0,
    '中華料理シャン': 0,
    'sandish': 0,
    'モスバーガー美野島店':0,
    'そみ': 0,
    'COMFORT Stand HAKATA': 0
}

# 店名と画像を定義
food_items = {
    'カレー': {'image': 'curry.jpg', 'store': 'カレー'},
    'ラーメン': {'image': 'ramen.jpg', 'store': 'ラーメン'},
    'コンビニ': {'image': 'convenience_store.jpg', 'store': 'コンビニ'},
    'スーパー': {'image': 'supermarket.jpg', 'store': 'スーパー'},
    'うどん平': {'image': 'udon.jpg', 'store': 'うどん平'},
    'マカロニキッチン': {'image': 'makaroni.jpg', 'store': 'マカロニキッチン'},
    'キッチングローリ':{'image': 'guro-ri.jpg', 'store': 'キッチングローリ'},
    '九一郎':{'image': 'kuitirou.jpg', 'store': '九一郎'},
    '食処 やま利':{'image': 'yamari.jpg', 'store': '食処 やま利'},
    '元祖ビックリ亭':{'image': 'bikkuritei.jpg', 'store': '元祖ビックリ亭'},
    'かどや食堂':{'image': 'kadoyasyokudou.jpg', 'store': 'かどや食堂'},
    'みかちゃん家のいろどりごはん':{'image': 'mikachan.jpg', 'store': 'みかちゃん家のいろどりごはん'},
    'ロイヤルホスト':{'image': 'royal_host.jpg', 'store': 'ロイヤルホスト'},
    'キッシュとパンのジョー': {'image': 'kissyu.jpg', 'store': 'キッシュとパンのジョー'},
    'Home Kitchen COTALO': {'image': 'COTALO.jpg', 'store': 'Home Kitchen COTALO'},
    '中華料理シャン':{'image': 'syan.jpg', 'store': '中華料理シャン'},
    'sandish':{'image': 'sandish.jpg', 'store': 'sandish'},
    'モスバーガー美野島店':{'image': 'mos.jpg', 'store': 'モスバーガー美野島店'},
    'そみ': {'image': 'somi.jpg', 'store': 'そみ'},
    'COMFORT Stand HAKATA':{'image': 'comfort.jpg', 'store': 'COMFORT Stand HAKATA'}
}

# スタンプコンプリートまでの閾値
target_stamps = 5

# 音声ファイルを絶対パスに置き換えることでエラー回避
sound_file = os.path.abspath("explosive.mp3")

# エクセルファイル作成時の名前
excel_file = "visit_count.xlsx"

# エクセルシート内への書き込み
def write_to_excel():
    current_month = datetime.now().strftime("%Y年%m月")
    try:
        workbook = load_workbook(excel_file)
    except FileNotFoundError:
        workbook = Workbook()    # 当月のシートがなかったら作成

    if current_month not in workbook.sheetnames:
        workbook.create_sheet(current_month)

    sheet = workbook[current_month]
    if sheet.max_row == 1 and sheet.cell(row=1, column=1).value is None:
        sheet.cell(row=1, column=1, value="店舗")
        sheet.cell(row=1, column=2, value="訪問回数")

    for store, visits in food_stamps.items():
        store_found = False
        for row in range(2, sheet.max_row + 1):
            if sheet.cell(row=row, column=1).value == store:
                sheet.cell(row=row, column=2, value=visits)
                store_found = True
                break
        if not store_found:
            last_row = sheet.max_row + 1
            sheet.cell(row=last_row, column=1, value=store)
            sheet.cell(row=last_row, column=2, value=visits)

    workbook.save(excel_file)

# エクセルの読み込み
def read_from_excel():
    current_month = datetime.now().strftime("%Y年%m月")
    try:
        workbook = load_workbook(excel_file)
        if current_month in workbook.sheetnames:
            sheet = workbook[current_month]
            for row in range(2, sheet.max_row + 1):
                store = sheet.cell(row=row, column=1).value
                visits = sheet.cell(row=row, column=2).value
                if store in food_stamps:
                    food_stamps[store] = visits
    except FileNotFoundError:
        pass    # ファイルがまだなくても正常に動くよう工夫

read_from_excel()

# アプリのラベルと各辞書をリンク
def dispLabel():
    selected_food = random.choice(list(food_items.keys()))
    food_stamps[selected_food] += 1    # 選ばれた店舗は+1カウント

    lbl.configure(text=f"今日のランチはこれ！！")
    store_name_label.configure(text=f"{food_items[selected_food]['store']}")
    
    disPhoto(food_items[selected_food]["image"])
    update_stamp_label(selected_food)

    write_to_excel()

    if food_stamps[selected_food] == target_stamps:
        complete_label.configure(text=f"🎉{selected_food}のスタンプラリーコンプリート！ 何もでないよ！笑🎉")
        threading.Thread(target=play_sound).start()    # フリーズ回避

    map_button.configure(command=lambda:open_map(food_items[selected_food]["store"]))
    map_button.pack()

# 選ばれた店舗のカウント数をここで更新
def update_stamp_label(selected_food):
    text = f"{food_stamps[selected_food]} 回行ったよ！"
    stamp_label.configure(text=text)

# アプリ内での画像の出し方
def disPhoto(path):
    newImage = PIL.Image.open(path).resize((400, 300))
    imageData = PIL.ImageTk.PhotoImage(newImage)
    imageLabel.configure(image=imageData)
    imageLabel.image = imageData

# アプリ自体の背景画面を設定
def set_background(canvas, image_path):
    bg_image = Image.open(image_path)
    bg_image = bg_image.resize((600, 600))
    bg_photo = ImageTk.PhotoImage(bg_image)
    
    canvas.create_image(0, 0, image=bg_photo, anchor=tk.NW)
    canvas.image = bg_photo

# 音声設定
def play_sound():
    playsound(sound_file)

# 地図の設定
def open_map(store_name):
    search_query = f"https://www.google.com/maps/search/{store_name}"
    webbrowser.open(search_query)


# アプリ自体の設定
root = tk.Tk()
root.geometry('600x600')

canvas = tk.Canvas(root, width=600, height=600)
canvas.pack(fill="both", expand=True)

background_image_path = "background.jpg"
set_background(canvas, background_image_path)

lbl = tk.Label(canvas, text="今日のランチはこれ！！")
lbl.pack()

store_name_label = tk.Label(canvas, text="")
store_name_label.pack()

stamp_label = tk.Label(canvas, text="")
stamp_label.pack()

btn = tk.Button(canvas, text="デデーン！！", command=dispLabel)
btn.pack()

imageLabel = tk.Label(canvas)
imageLabel.pack()

map_button = tk.Button(canvas, text="早く食べに行くウッキー🍌🐵", command=lambda:None)
complete_label = tk.Label(canvas, text="")
complete_label.pack()

tk.mainloop()